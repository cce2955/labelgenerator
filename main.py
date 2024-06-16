import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
import os

def get_user_input():
    data = []
    for row in range(22):  # Adjust for the new number of rows
        row_data = []
        for col in range(8):
            user_input = input(f"Enter input for cell ({row+1}, {col+1}) (type 'exit' to stop): ")
            if user_input.lower() == 'exit':
                while len(row_data) < 8:  # Fill the remaining cells in the row with empty strings
                    row_data.append('')
                data.append(row_data)
                while len(data) < 22:  # Fill the remaining rows with empty strings
                    data.append([''] * 8)
                return data
            row_data.append(user_input)
        data.append(row_data)
    return data

def find_next_available_cell(ws):
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col).value is None:
                return row, col
    return ws.max_row + 1, 1  # If no empty cell is found, start at the next row

def create_excel_sheet(data, append_mode=False):
    if append_mode and os.path.exists("user_input_grid.xlsx"):
        wb = openpyxl.load_workbook("user_input_grid.xlsx")
        ws = wb.active
        start_row, start_col = find_next_available_cell(ws)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        start_row, start_col = 1, 1

    # Set the column width and row height to approximate 1 inch wide and 0.5 inch high cells
    one_inch_points = 72  # 1 inch = 72 points
    col_width = 10  # Width in Excel units, adjust if necessary
    row_height = one_inch_points / 2  # Half inch in points

    for col in range(1, 9):  # Excel columns start from 1
        ws.column_dimensions[get_column_letter(col)].width = col_width

    for row in range(1, 23):  # Adjust for the new number of rows
        ws.row_dimensions[row].height = row_height

    # Define a border style for the cutting guide
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    max_font_size = 14
    min_font_size = 6

    # Fill the cells with data and apply borders and dynamic font size
    row, col = start_row, start_col
    for row_data in data:
        for value in row_data:
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            
            # Calculate appropriate font size based on length of the input
            if len(value) <= 10:
                font_size = max_font_size
            else:
                font_size = max(max_font_size - len(value) // 2, min_font_size)
            
            if font_size < min_font_size:
                cell.value = value[:int((max_font_size - min_font_size) * 2)]  # Truncate to fit within the limits
                font_size = min_font_size

            cell.font = Font(size=font_size)
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

            col += 1
            if col > 8:
                col = 1
                row += 1

    # Set print options
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.print_area = 'A1:H22'  # Ensure the print area covers the new grid size
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # Save the workbook
    wb.save("user_input_grid.xlsx")
    print("Excel sheet created as 'user_input_grid.xlsx'.")

def main():
    user_choice = input("Do you want to append to the existing file or create a new one? (A for append, N for new): ").strip().upper()
    if user_choice == 'A':
        append_mode = True
    elif user_choice == 'N':
        append_mode = False
    else:
        print("Invalid choice. Please enter 'A' to append or 'N' to create a new file.")
        return

    print("You will be prompted to enter inputs for an 8x22 grid (1 inch wide, 0.5 inch high cells).")
    data = get_user_input()
    create_excel_sheet(data, append_mode)

if __name__ == "__main__":
    main()
